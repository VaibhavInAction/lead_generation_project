"""Offline tests for the intent-lead exporters (README §18, §23)."""

from __future__ import annotations

import csv
from datetime import UTC, datetime

import pytest
from openpyxl import load_workbook

from leadforge.exports.intent import INTENT_COLUMNS, write_intent_leads

LONG_URL = (
    "https://www.linkedin.com/posts/paul-mccarron-79785053_were-hiring-"
    "activity-7480596109783048192-ODcs?utm_source=share&utm_medium=member_desktop"
)

ROWS: list[dict[str, object]] = [
    {
        "author_name": "Paul Mccarron",
        "lead_score": 0,  # a hiring post -> forced to the bottom (Phase 9)
        "category": "job_posting",
        "need_text": "We're hiring! We're looking for a Marketing Manager.",
        "post_url": LONG_URL,
        "posted_at": datetime(2026, 7, 8, 12, 18, tzinfo=UTC),
        "platform": "linkedin_public",
    },
    {
        "author_name": "Sarah Udaipurwala",
        "lead_score": 88,
        "category": "client_lead",
        "need_text": "Looking for a marketing/growth advisor.",
        "post_url": "https://www.linkedin.com/posts/sarah-udaipurwala_activity-1",
        "posted_at": None,  # genuinely absent -> empty cell
        "platform": "linkedin_public",
    },
]


class TestWriteCsv:
    def test_writes_header_values_and_full_url(self, tmp_path) -> None:
        path = tmp_path / "out.csv"
        write_intent_leads(ROWS, path, "csv")

        with path.open(encoding="utf-8-sig", newline="") as handle:
            records = list(csv.reader(handle))

        assert records[0] == list(INTENT_COLUMNS)
        row = dict(zip(INTENT_COLUMNS, records[1], strict=True))
        assert row["author_name"] == "Paul Mccarron"
        assert row["lead_score"] == "0"
        assert row["category"] == "job_posting"
        assert row["need_text"] == "We're hiring! We're looking for a Marketing Manager."
        assert row["post_url"] == LONG_URL  # full, un-truncated
        assert row["posted_at"] == "2026-07-08T12:18:00+00:00"
        assert row["platform"] == "linkedin_public"
        # Missing posted_at is an empty cell, not the string "None".
        assert records[2][INTENT_COLUMNS.index("posted_at")] == ""

    def test_uses_utf8_bom_for_excel(self, tmp_path) -> None:
        path = tmp_path / "out.csv"
        write_intent_leads(ROWS, path, "csv")
        assert path.read_bytes().startswith(b"\xef\xbb\xbf")  # UTF-8 BOM


class TestWriteXlsx:
    def test_writes_rows_with_clickable_post_url(self, tmp_path) -> None:
        path = tmp_path / "out.xlsx"
        write_intent_leads(ROWS, path, "xlsx")

        workbook = load_workbook(path)
        sheet = workbook.active

        assert [cell.value for cell in sheet[1]] == list(INTENT_COLUMNS)
        url_col = INTENT_COLUMNS.index("post_url") + 1
        url_cell = sheet.cell(row=2, column=url_col)
        assert url_cell.value == LONG_URL  # full URL as the cell text
        assert url_cell.hyperlink is not None
        assert url_cell.hyperlink.target == LONG_URL  # and a clickable hyperlink
        sheet_freeze = sheet.freeze_panes
        assert sheet_freeze == "A2"  # header row frozen


def test_unsupported_format_raises(tmp_path) -> None:
    with pytest.raises(ValueError, match="unsupported export format"):
        write_intent_leads(ROWS, tmp_path / "out.txt", "txt")
